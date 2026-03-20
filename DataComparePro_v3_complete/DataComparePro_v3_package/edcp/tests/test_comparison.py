# -*- coding: utf-8 -*-
"""
tests/test_comparison.py
─────────────────────────
TC-01  CSV vs CSV – matched passed, failed, missing, extra
TC-02  TXT vs TXT – pipe-delimited auto-detected
TC-03  XLSX vs XLSX
TC-04  CSV vs XLSX cross-format
TC-05  TXT vs CSV  cross-format
TC-06  Duplicate key detection + _SEQ_ alignment
TC-07  Ignore fields excluded from all results
TC-08  Missing and extra records in Count Differences
TC-09  Pass Comparison respects include_pass_report flag
TC-10  Schema differences detected (missing + extra columns)
TC-11  All workbook sheets present
TC-12  Selective pipeline (comparison + DQ only)
TC-13  Context isolation across concurrent runs
TC-14  Empty file does not crash
TC-15  Row-order fallback when no keys defined
"""
from __future__ import annotations
import sys, os, tempfile, unittest
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from pathlib import Path
import pandas as pd
from data_compare.context.run_context import make_context
from data_compare.registry.capability_registry import CapabilityRegistry
from data_compare.reporting.report_builder import build_report

_PROD = (
    "TradeID,Portfolio,Price,Quantity,Trader,IgnoreMe\n"
    "T001,EQ,100.001,500,Alice,X\nT002,FI,50.500,200,Bob,Y\n"
    "T003,EQ,75.000,300,Carol,X\nT003,EQ,75.000,300,Carol,X\n"  # dup
    "T004,FI,200.000,100,Dave,Y\nT005,EQ,999.000,50,Eve,X\n"
)
_DEV = (
    "TradeID,Portfolio,Price,Quantity,TraderName,IgnoreMe\n"
    "T001,EQ,100.003,500,Alice,A\nT002,FI,50.500,999,Bob,B\n"
    "T003,EQ,75.000,300,Carol,A\nT004,FI,200.000,100,Dave,A\n"
    "T006,EQ,123.000,777,Frank,B\n"
)

def _caps(**overrides):
    base = {c: True for c in ["comparison","tolerance","duplicate","schema","data_quality","audit","alerts"]}
    base["plugins"] = False; base["parquet"] = False
    base.update(overrides)
    return base

def _run(prod_path, dev_path, prod_name, dev_name, tmp, caps=None, tol_map=None, keys=None, ignore=None, include_pass=True):
    from data_compare.loaders.file_loader import load_any_to_csv
    conv = tmp / "conv"; conv.mkdir(exist_ok=True)
    pc = load_any_to_csv(prod_path, conv); dc = load_any_to_csv(dev_path, conv)
    ctx = make_context(
        prod_csv_path=pc, dev_csv_path=dc, prod_name=prod_name, dev_name=dev_name,
        result_name="test", report_root=tmp/"reports",
        keys=keys or ["TradeID","Portfolio"], ignore_fields=ignore or ["IgnoreMe"],
        tol_map=tol_map or {("prod.csv","dev.csv","Price"):2},
        capabilities_cfg=caps or _caps(),
        alert_rules=[{"metric":"matched_failed","operator":">","threshold":0,"level":"WARNING","message":"diffs"},
                     {"metric":"missing_dev","operator":">","threshold":0,"level":"INFO","message":"miss"},
                     {"metric":"extra_dev","operator":">","threshold":0,"level":"INFO","message":"extra"},
                     {"metric":"dup_count_prod","operator":">","threshold":0,"level":"INFO","message":"dups"}],
        include_pass_report=include_pass,
    )
    return CapabilityRegistry().run_pipeline(ctx)

class TC01_CSV_vs_CSV(unittest.TestCase):
    def setUp(self):
        self._d = tempfile.TemporaryDirectory(); self.t = Path(self._d.name)
        (self.t/"prod.csv").write_text(_PROD); (self.t/"dev.csv").write_text(_DEV)
    def tearDown(self): self._d.cleanup()
    def test_TC01a_matched_passed(self):
        ctx = _run(self.t/"prod.csv", self.t/"dev.csv", "prod.csv", "dev.csv", self.t)
        self.assertGreaterEqual(ctx["results"]["matched_passed"], 2)
    def test_TC01b_matched_failed(self):
        ctx = _run(self.t/"prod.csv", self.t/"dev.csv", "prod.csv", "dev.csv", self.t)
        self.assertEqual(ctx["results"]["matched_failed"], 1)
    def test_TC01c_missing_in_dev(self):
        ctx = _run(self.t/"prod.csv", self.t/"dev.csv", "prod.csv", "dev.csv", self.t)
        self.assertGreaterEqual(len(ctx["only_in_prod"]), 1)
    def test_TC01d_extra_in_dev(self):
        ctx = _run(self.t/"prod.csv", self.t/"dev.csv", "prod.csv", "dev.csv", self.t)
        self.assertEqual(len(ctx["only_in_dev"]), 1)
    def test_TC01e_dup_detected(self):
        ctx = _run(self.t/"prod.csv", self.t/"dev.csv", "prod.csv", "dev.csv", self.t)
        self.assertEqual(ctx["dup_count_prod"], 1)
    def test_TC01f_tolerance_active(self):
        ctx = _run(self.t/"prod.csv", self.t/"dev.csv", "prod.csv", "dev.csv", self.t)
        self.assertIn("Price", ctx["tol_for_pair"])
    def test_TC01g_schema_missing(self):
        ctx = _run(self.t/"prod.csv", self.t/"dev.csv", "prod.csv", "dev.csv", self.t)
        sd = ctx["schema_diff_df"]; self.assertIsNotNone(sd)
        self.assertIn("Trader", sd[sd["Side"]=="Missing in Dev"]["Field"].tolist())
    def test_TC01h_schema_extra(self):
        ctx = _run(self.t/"prod.csv", self.t/"dev.csv", "prod.csv", "dev.csv", self.t)
        self.assertIn("TraderName", ctx["schema_diff_df"][ctx["schema_diff_df"]["Side"]=="Extra in Dev"]["Field"].tolist())
    def test_TC01i_ignore_field(self):
        ctx = _run(self.t/"prod.csv", self.t/"dev.csv", "prod.csv", "dev.csv", self.t)
        self.assertIn("IgnoreMe", ctx["ignore_set"])
        gd = ctx["results"]["grouped_differences"]
        if gd is not None and not gd.empty:
            self.assertNotIn("IgnoreMe", gd["Column"].tolist())
    def test_TC01j_report_written(self):
        ctx = _run(self.t/"prod.csv", self.t/"dev.csv", "prod.csv", "dev.csv", self.t)
        out = build_report(ctx); self.assertTrue(out.exists())

class TC02_TXT_vs_TXT(unittest.TestCase):
    def setUp(self):
        self._d = tempfile.TemporaryDirectory(); self.t = Path(self._d.name)
        (self.t/"prod.txt").write_text(_PROD.replace(",","|"))
        (self.t/"dev.txt").write_text(_DEV.replace(",","|"))
    def tearDown(self): self._d.cleanup()
    def _ctx(self): return _run(self.t/"prod.txt", self.t/"dev.txt", "prod.txt","dev.txt", self.t,
                                tol_map={("prod.txt","dev.txt","Price"):2})
    def test_TC02a_failed(self): self.assertEqual(self._ctx()["results"]["matched_failed"], 1)
    def test_TC02b_dup(self): self.assertEqual(self._ctx()["dup_count_prod"], 1)
    def test_TC02c_schema(self): self.assertFalse(self._ctx()["schema_diff_df"].empty)
    def test_TC02d_report(self): self.assertTrue(build_report(self._ctx()).exists())

class TC03_XLSX_vs_XLSX(unittest.TestCase):
    def setUp(self):
        self._d = tempfile.TemporaryDirectory(); self.t = Path(self._d.name)
        pd.read_csv(self.t/"prod.csv") if False else None
        prod_df = pd.read_csv(__import__("io").StringIO(_PROD))
        dev_df  = pd.read_csv(__import__("io").StringIO(_DEV))
        prod_df.to_excel(self.t/"prod.xlsx", index=False)
        dev_df.to_excel(self.t/"dev.xlsx",  index=False)
    def tearDown(self): self._d.cleanup()
    def _ctx(self): return _run(self.t/"prod.xlsx",self.t/"dev.xlsx","prod.xlsx","dev.xlsx",self.t,
                                tol_map={("prod.xlsx","dev.xlsx","Price"):2})
    def test_TC03a_rows(self):
        ctx = self._ctx()
        self.assertEqual(len(ctx["prod_df"]), 6); self.assertEqual(len(ctx["dev_df"]), 5)
    def test_TC03b_failed(self): self.assertEqual(self._ctx()["results"]["matched_failed"], 1)
    def test_TC03c_report(self): self.assertTrue(build_report(self._ctx()).exists())

class TC04_CSV_vs_XLSX(unittest.TestCase):
    def setUp(self):
        self._d = tempfile.TemporaryDirectory(); self.t = Path(self._d.name)
        import io
        (self.t/"prod.csv").write_text(_PROD)
        pd.read_csv(io.StringIO(_DEV)).to_excel(self.t/"dev.xlsx", index=False)
    def tearDown(self): self._d.cleanup()
    def test_TC04_failed_and_dup(self):
        ctx = _run(self.t/"prod.csv",self.t/"dev.xlsx","prod.csv","dev.xlsx",self.t,
                   tol_map={("prod.csv","dev.xlsx","Price"):2})
        self.assertEqual(ctx["results"]["matched_failed"], 1)
        self.assertEqual(ctx["dup_count_prod"], 1)

class TC05_TXT_vs_CSV(unittest.TestCase):
    def setUp(self):
        self._d = tempfile.TemporaryDirectory(); self.t = Path(self._d.name)
        (self.t/"prod.txt").write_text(_PROD.replace(",","|"))
        (self.t/"dev.csv").write_text(_DEV)
    def tearDown(self): self._d.cleanup()
    def test_TC05_failed_and_dup(self):
        ctx = _run(self.t/"prod.txt",self.t/"dev.csv","prod.txt","dev.csv",self.t,
                   tol_map={("prod.txt","dev.csv","Price"):2})
        self.assertEqual(ctx["results"]["matched_failed"], 1)
        self.assertEqual(ctx["dup_count_prod"], 1)

class TC06_Duplicates(unittest.TestCase):
    def setUp(self):
        self._d = tempfile.TemporaryDirectory(); self.t = Path(self._d.name)
        (self.t/"p.csv").write_text("ID,V\nA,1\nA,2\nB,3\n")
        (self.t/"d.csv").write_text("ID,V\nA,1\nA,2\nB,3\n")
    def tearDown(self): self._d.cleanup()
    def test_TC06a_dup_both(self):
        ctx = _run(self.t/"p.csv",self.t/"d.csv","p.csv","d.csv",self.t,
                   keys=["ID"],ignore=[],tol_map={},
                   caps=_caps(tolerance=False,schema=False,data_quality=False,audit=False,alerts=False))
        self.assertEqual(ctx["dup_count_prod"], 1); self.assertEqual(ctx["dup_count_dev"], 1)
        self.assertTrue(ctx["using_seq"])
    def test_TC06b_dup_identical_all_pass(self):
        ctx = _run(self.t/"p.csv",self.t/"d.csv","p.csv","d.csv",self.t,
                   keys=["ID"],ignore=[],tol_map={},
                   caps=_caps(tolerance=False,schema=False,data_quality=False,audit=False,alerts=False))
        self.assertEqual(ctx["results"]["matched_failed"], 0)

class TC07_IgnoreFields(unittest.TestCase):
    def setUp(self):
        self._d = tempfile.TemporaryDirectory(); self.t = Path(self._d.name)
        (self.t/"p.csv").write_text("ID,Val,Secret\n1,A,XXXX\n2,B,YYYY\n")
        (self.t/"d.csv").write_text("ID,Val,Secret\n1,A,DIFF\n2,B,DIFF2\n")
    def tearDown(self): self._d.cleanup()
    def test_TC07_secret_excluded(self):
        ctx = _run(self.t/"p.csv",self.t/"d.csv","p.csv","d.csv",self.t,
                   keys=["ID"],ignore=["Secret"],tol_map={},
                   caps=_caps(tolerance=False,schema=False,data_quality=False,audit=False,alerts=False))
        self.assertIn("Secret", ctx["ignore_set"])
        self.assertEqual(ctx["results"]["matched_passed"], 2)

class TC08_MissingExtra(unittest.TestCase):
    def setUp(self):
        self._d = tempfile.TemporaryDirectory(); self.t = Path(self._d.name)
        (self.t/"p.csv").write_text("ID,V\n1,A\n2,B\n3,C\n")
        (self.t/"d.csv").write_text("ID,V\n1,A\n4,D\n")
    def tearDown(self): self._d.cleanup()
    def test_TC08_missing_and_extra(self):
        ctx = _run(self.t/"p.csv",self.t/"d.csv","p.csv","d.csv",self.t,
                   keys=["ID"],ignore=[],tol_map={},
                   caps=_caps(tolerance=False,schema=False,data_quality=False,audit=False,alerts=False))
        self.assertEqual(len(ctx["only_in_prod"]), 2)   # 2,3
        self.assertEqual(len(ctx["only_in_dev"]),  1)   # 4
        me = ctx["results"]["missing_extra"]
        self.assertFalse(me.empty)
        self.assertIn("Missing in Dev", me["Side"].tolist())

class TC09_PassReport(unittest.TestCase):
    def setUp(self):
        self._d = tempfile.TemporaryDirectory(); self.t = Path(self._d.name)
        (self.t/"p.csv").write_text("ID,V\n1,A\n2,B\n"); (self.t/"d.csv").write_text("ID,V\n1,A\n2,X\n")
    def tearDown(self): self._d.cleanup()
    def _ctx(self, include): return _run(self.t/"p.csv",self.t/"d.csv","p.csv","d.csv",self.t,
                                         keys=["ID"],ignore=[],tol_map={},include_pass=include,
                                         caps=_caps(tolerance=False,schema=False,data_quality=False,audit=False,alerts=False))
    def test_TC09a_include_true(self): self.assertGreaterEqual(len(self._ctx(True)["results"]["pass_report"]), 2)
    def test_TC09b_include_false(self):
        pr = self._ctx(False)["results"]["pass_report"]
        if "Info" not in pr.columns: self.assertEqual(len(pr), 1)

class TC10_Schema(unittest.TestCase):
    def setUp(self):
        self._d = tempfile.TemporaryDirectory(); self.t = Path(self._d.name)
        (self.t/"p.csv").write_text("ID,A,B\n1,x,y\n"); (self.t/"d.csv").write_text("ID,A,C\n1,x,z\n")
    def tearDown(self): self._d.cleanup()
    def test_TC10_schema_diff(self):
        ctx = _run(self.t/"p.csv",self.t/"d.csv","p.csv","d.csv",self.t,
                   keys=["ID"],ignore=[],tol_map={},
                   caps=_caps(tolerance=False,duplicate=False,data_quality=False,audit=False,alerts=False))
        sd = ctx["schema_diff_df"]
        self.assertIn("B", sd[sd["Side"]=="Missing in Dev"]["Field"].tolist())
        self.assertIn("C", sd[sd["Side"]=="Extra in Dev"]["Field"].tolist())

class TC11_WorkbookSheets(unittest.TestCase):
    def setUp(self):
        self._d = tempfile.TemporaryDirectory(); self.t = Path(self._d.name)
        (self.t/"p.csv").write_text("ID,V\n1,A\n"); (self.t/"d.csv").write_text("ID,V\n1,B\n")
    def tearDown(self): self._d.cleanup()
    def test_TC11_all_sheets(self):
        from openpyxl import load_workbook
        ctx = _run(self.t/"p.csv",self.t/"d.csv","p.csv","d.csv",self.t,
                   keys=["ID"],ignore=[],tol_map={},caps=_caps(alerts=False))
        wb = load_workbook(build_report(ctx))
        for s in ["Grouped Differences","Count Differences","Pass Comparison",
                  "Duplicates Records","Schema Differences","Settings",
                  "Data Quality Report","Audit Trail"]:
            self.assertIn(s, wb.sheetnames, f"Missing: {s}")

class TC12_SelectivePipeline(unittest.TestCase):
    def setUp(self):
        self._d = tempfile.TemporaryDirectory(); self.t = Path(self._d.name)
        (self.t/"p.csv").write_text("ID,V\n1,A\n2,B\n"); (self.t/"d.csv").write_text("ID,V\n1,A\n2,X\n")
    def tearDown(self): self._d.cleanup()
    def test_TC12_schema_none_when_disabled(self):
        ctx = _run(self.t/"p.csv",self.t/"d.csv","p.csv","d.csv",self.t,
                   keys=["ID"],ignore=[],tol_map={},
                   caps=_caps(tolerance=False,duplicate=False,schema=False,audit=False,alerts=False))
        self.assertIsNone(ctx["schema_diff_df"])
        self.assertIsNotNone(ctx["metrics"]["data_quality_report"])

class TC13_ContextIsolation(unittest.TestCase):
    def test_TC13_isolated(self):
        a = make_context(result_name="A"); b = make_context(result_name="B")
        a["results"]["matched_failed"] = 999
        self.assertEqual(b["results"]["matched_failed"], 0)
        self.assertNotEqual(a["audit"]["run_id"], b["audit"]["run_id"])

class TC14_EmptyFile(unittest.TestCase):
    def setUp(self):
        self._d = tempfile.TemporaryDirectory(); self.t = Path(self._d.name)
        (self.t/"p.csv").write_text("ID,V\n"); (self.t/"d.csv").write_text("ID,V\n1,A\n")
    def tearDown(self): self._d.cleanup()
    def test_TC14_no_crash(self):
        ctx = _run(self.t/"p.csv",self.t/"d.csv","p.csv","d.csv",self.t,
                   keys=["ID"],ignore=[],tol_map={},
                   caps=_caps(tolerance=False,duplicate=False,schema=False,data_quality=False,audit=False,alerts=False))
        self.assertIsNotNone(ctx)

class TC15_RowOrderFallback(unittest.TestCase):
    def setUp(self):
        self._d = tempfile.TemporaryDirectory(); self.t = Path(self._d.name)
        (self.t/"p.csv").write_text("Val\nA\nB\nC\n"); (self.t/"d.csv").write_text("Val\nA\nX\nC\n")
    def tearDown(self): self._d.cleanup()
    def test_TC15_no_keys(self):
        ctx = _run(self.t/"p.csv",self.t/"d.csv","p.csv","d.csv",self.t,
                   keys=[],ignore=[],tol_map={},
                   caps=_caps(tolerance=False,duplicate=False,schema=False,data_quality=False,audit=False,alerts=False))
        self.assertEqual(ctx["results"]["matched_failed"], 1)
        self.assertEqual(ctx["results"]["matched_passed"], 2)

if __name__ == "__main__":
    import unittest; unittest.main(verbosity=2)
