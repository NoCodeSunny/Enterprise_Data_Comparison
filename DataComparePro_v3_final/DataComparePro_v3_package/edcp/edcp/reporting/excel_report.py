# -*- coding: utf-8 -*-
"""
data_compare.reporting.excel_report
──────────────────────────────────────
Excel workbook generation.

Writes all comparison sheets to a single openpyxl workbook in one atomic
save (no double-open, no intermediate read-back).

Features:
  [F-38]  Light-blue bold header row on every sheet
  [F-39]  Auto-width columns (capped at 60 characters)
  [F-40]  Red fill for "Fail" cells / green fill for "Pass" cells in Pass Comparison
  [F-41]  One-pass write – no double-open
  [F-46]  Legacy build_html_summary_table() retained (in html_report.py)
"""

from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from data_compare.utils.logger import get_logger

logger = get_logger(__name__)

# ── style constants ───────────────────────────────────────────────────────────
_HEADER_FILL  = PatternFill("solid", fgColor="D9E1F2")
_HEADER_FONT  = Font(bold=True)
_HEADER_ALIGN = Alignment(vertical="center", wrap_text=False)
_RED_FILL     = PatternFill("solid", fgColor="FFC7CE")
_GREEN_FILL   = PatternFill("solid", fgColor="C6EFCE")


def _calc_col_widths(ws, max_width: int = 60) -> Dict[int, float]:
    """Compute display widths from cell content; capped at *max_width*."""
    widths: Dict[int, float] = {}
    for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
        max_len = 0
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len  = max(max_len, cell_len)
            except Exception:
                pass
        widths[col_idx] = min(max_len + 2, max_width)
    return widths


def _apply_sheet_style(
    ws,
    compare_col_idxs: Optional[List[int]] = None,
) -> None:
    """
    Apply:
      - Bold light-blue header row
      - Auto-width columns
      - Red/green cell fill for Pass/Fail cells (Pass Comparison sheet only)
    """
    # Header row
    for cell in ws[1]:
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    # Column widths
    widths = _calc_col_widths(ws)
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Colour Pass / Fail cells
    if compare_col_idxs:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for ci in compare_col_idxs:
                cell = row[ci - 1]
                val  = str(cell.value).strip().lower() if cell.value else ""
                if val == "fail":
                    cell.fill = _RED_FILL
                elif val == "pass":
                    cell.fill = _GREEN_FILL


def write_workbook_once(
    sheets: Dict[str, pd.DataFrame],
    out_path: Path,
    pass_sheet_name: str = "Pass Comparison",
) -> None:
    """
    Write *sheets* to a single openpyxl workbook in one atomic save.

    Parameters
    ----------
    sheets : dict
        Ordered mapping of sheet_name → DataFrame.  Sheets are written in
        the dict's iteration order.
    out_path : Path
        Destination file path.  Must not exist (caller uses safe_output_path).
    pass_sheet_name : str
        Name of the sheet where Pass/Fail colouring should be applied.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(title=sheet_name)

        # Guard against None / empty DataFrames
        if df is None or df.empty:
            df = pd.DataFrame({"Info": ["No data"]})

        # Write header row
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=1, column=col_idx, value=str(col_name))

        # Write data rows
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, val in enumerate(row, start=1):
                ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=str(val) if val is not None else "",
                )

        # Identify CompareResult columns for colouring
        compare_idxs: Optional[List[int]] = None
        if sheet_name == pass_sheet_name:
            compare_idxs = [
                i for i, c in enumerate(df.columns, start=1)
                if str(c).startswith("CompareResult_")
            ]

        _apply_sheet_style(ws, compare_idxs)

    wb.save(out_path)
    logger.debug(f"  Workbook saved: {out_path}")
