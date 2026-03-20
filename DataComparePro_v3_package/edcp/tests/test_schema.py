# -*- coding: utf-8 -*-
"""
tests/test_schema.py
─────────────────────
TS-01  build_schema_diff_dataframe: missing column detected
TS-02  build_schema_diff_dataframe: extra column detected
TS-03  build_schema_diff_dataframe: no diff returns empty DataFrame
TS-04  build_schema_diff_dataframe: excludes key columns
TS-05  build_schema_diff_dataframe: excludes ignore_fields
TS-06  annotate_schema_counts adds Count column
TS-07  SchemaCapability standalone sets schema_diff_df
TS-08  SchemaCapability disabled leaves schema_diff_df None
TS-09  Pipeline schema: Trader missing, TraderName extra
TS-10  Pipeline schema: multiple missing and extra columns
TS-11  Schema diff writes Schema Differences sheet in workbook
TS-12  Schema diff with all columns matching returns empty diff
TS-13  Schema excludes _SEQ_ from diff when alignment active
"""
from __future__ import annotations
import sys, os, tempfile, unittest
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from pathlib import Path
import pandas as pd
from data_compare.comparator.schema import (
    build_schema_diff_dataframe, annotate_schema_counts
)
from data_compare.context.run_context import make_context
from data_compare.registry.capability_registry import CapabilityRegistry

def _caps_schema():
    return {"parquet":False,"comparison":True,"tolerance":False,"duplicate":False,
            "schema":True,"data_quality":False,"audit":False,"alerts":False,"plugins":False}

class TS01_Missing(unittest.TestCase):
    def test_TS01(self):
        df = build_schema_diff_dataframe(["ID","A","B"],["ID","A"],set())
        self.assertIn("B", df[df["Side"]=="Missing in Dev"]["Field"].tolist())

class TS02_Extra(unittest.TestCase):
    def test_TS02(self):
        df = build_schema_diff_dataframe(["ID","A"],["ID","A","C"],set())
        self.assertIn("C", df[df["Side"]=="Extra in Dev"]["Field"].tolist())

class TS03_NoDiff(unittest.TestCase):
    def test_TS03(self):
        df = build_schema_diff_dataframe(["ID","A","B"],["ID","A","B"],set())
        self.assertTrue(df.empty)

class TS04_ExcludesKeys(unittest.TestCase):
    def test_TS04(self):
        df = build_schema_diff_dataframe(["ID","A"],["ID","B"],{"ID"})
        fields = df["Field"].tolist()
        self.assertNotIn("ID", fields)

class TS05_ExcludesIgnore(unittest.TestCase):
    def test_TS05(self):
        df = build_schema_diff_dataframe(["ID","Secret","A"],["ID","A"],{"Secret"})
        self.assertNotIn("Secret", df["Field"].tolist())

class TS06_AnnotateCounts(unittest.TestCase):
    def test_TS06(self):
        diff_df = pd.DataFrame({"Field":["B"],"Side":["Missing in Dev"]})
        prod_common = pd.DataFrame({"B":["x","y"]})
        dev_common  = pd.DataFrame({"A":["a","b"]})
        result = annotate_schema_counts(diff_df, prod_common, dev_common)
        self.assertIn("Count", result.columns)

class TS07_Standalone(unittest.TestCase):
    def test_TS07(self):
        from data_compare.capabilities.schema.schema_capability import SchemaCapability
        ctx = make_context(capabilities_cfg={"schema":True})
        ctx["prod_df"] = pd.DataFrame({"ID":["1"],"A":["x"],"B":["y"]})
        ctx["dev_df"]  = pd.DataFrame({"ID":["1"],"A":["x"],"C":["z"]})
        ctx = SchemaCapability().run(ctx)
        self.assertIsNotNone(ctx["schema_diff_df"])
        self.assertIn("B", ctx["schema_diff_df"][ctx["schema_diff_df"]["Side"]=="Missing in Dev"]["Field"].tolist())

class TS08_Disabled(unittest.TestCase):
    def test_TS08(self):
        from data_compare.capabilities.schema.schema_capability import SchemaCapability
        ctx = make_context(capabilities_cfg={"schema":False})
        ctx = SchemaCapability().run(ctx)
        self.assertIsNone(ctx["schema_diff_df"])

class TS09_Pipeline_SchemaRename(unittest.TestCase):
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TS09(self):
        from data_compare.loaders.file_loader import load_any_to_csv
        (self.t/"p.csv").write_text("ID,Price,Trader\nT001,100,Alice\n")
        (self.t/"d.csv").write_text("ID,Price,TraderName\nT001,100,Alice\n")
        conv=self.t/"c"; conv.mkdir()
        pc=load_any_to_csv(self.t/"p.csv",conv); dc=load_any_to_csv(self.t/"d.csv",conv)
        ctx = make_context(prod_csv_path=pc,dev_csv_path=dc,prod_name="p.csv",dev_name="d.csv",
            result_name="t",report_root=self.t,keys=["ID"],capabilities_cfg=_caps_schema())
        ctx = CapabilityRegistry().run_pipeline(ctx)
        sd = ctx["schema_diff_df"]; self.assertIsNotNone(sd)
        self.assertIn("Trader",     sd[sd["Side"]=="Missing in Dev"]["Field"].tolist())
        self.assertIn("TraderName", sd[sd["Side"]=="Extra in Dev"]["Field"].tolist())

class TS10_Pipeline_MultipleColumns(unittest.TestCase):
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TS10(self):
        from data_compare.loaders.file_loader import load_any_to_csv
        (self.t/"p.csv").write_text("ID,A,B,C\nT1,1,2,3\n")
        (self.t/"d.csv").write_text("ID,A,D,E\nT1,1,4,5\n")
        conv=self.t/"c"; conv.mkdir()
        pc=load_any_to_csv(self.t/"p.csv",conv); dc=load_any_to_csv(self.t/"d.csv",conv)
        ctx = make_context(prod_csv_path=pc,dev_csv_path=dc,prod_name="p.csv",dev_name="d.csv",
            result_name="t",report_root=self.t,keys=["ID"],capabilities_cfg=_caps_schema())
        ctx = CapabilityRegistry().run_pipeline(ctx)
        sd = ctx["schema_diff_df"]
        missing = sd[sd["Side"]=="Missing in Dev"]["Field"].tolist()
        extra   = sd[sd["Side"]=="Extra in Dev"]["Field"].tolist()
        self.assertIn("B", missing); self.assertIn("C", missing)
        self.assertIn("D", extra);   self.assertIn("E", extra)

class TS11_Workbook_Sheet(unittest.TestCase):
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TS11(self):
        from data_compare.loaders.file_loader import load_any_to_csv
        from data_compare.reporting.report_builder import build_report
        from openpyxl import load_workbook
        (self.t/"p.csv").write_text("ID,A,B\n1,x,y\n"); (self.t/"d.csv").write_text("ID,A,C\n1,x,z\n")
        conv=self.t/"c"; conv.mkdir()
        pc=load_any_to_csv(self.t/"p.csv",conv); dc=load_any_to_csv(self.t/"d.csv",conv)
        ctx = make_context(prod_csv_path=pc,dev_csv_path=dc,prod_name="p.csv",dev_name="d.csv",
            result_name="t",report_root=self.t,keys=["ID"],capabilities_cfg=_caps_schema())
        ctx = CapabilityRegistry().run_pipeline(ctx)
        wb = load_workbook(build_report(ctx))
        self.assertIn("Schema Differences", wb.sheetnames)

class TS12_NoSchemaDiff_Empty(unittest.TestCase):
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TS12(self):
        from data_compare.loaders.file_loader import load_any_to_csv
        (self.t/"p.csv").write_text("ID,V\n1,A\n"); (self.t/"d.csv").write_text("ID,V\n1,A\n")
        conv=self.t/"c"; conv.mkdir()
        pc=load_any_to_csv(self.t/"p.csv",conv); dc=load_any_to_csv(self.t/"d.csv",conv)
        ctx = make_context(prod_csv_path=pc,dev_csv_path=dc,prod_name="p.csv",dev_name="d.csv",
            result_name="t",report_root=self.t,keys=["ID"],capabilities_cfg=_caps_schema())
        ctx = CapabilityRegistry().run_pipeline(ctx)
        sd = ctx["schema_diff_df"]; self.assertTrue(sd is None or sd.empty)

class TS13_SEQ_ExcludedFromSchema(unittest.TestCase):
    def test_TS13(self):
        df = build_schema_diff_dataframe(
            ["ID","_SEQ_","A"],["ID","_SEQ_","A"],{"_SEQ_"}
        )
        if not df.empty:
            self.assertNotIn("_SEQ_", df["Field"].tolist())

if __name__ == "__main__":
    import unittest; unittest.main(verbosity=2)
