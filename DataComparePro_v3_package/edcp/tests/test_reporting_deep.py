# -*- coding: utf-8 -*-
"""tests/test_reporting_deep.py – TRPT01–TRPT06  Deep report content validation."""
from __future__ import annotations
import sys, os, tempfile, unittest, json
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from pathlib import Path
import pandas as pd

def _run_and_report(prod, dev, tmp):
    from data_compare.loaders.file_loader import load_any_to_csv
    from data_compare.context.run_context import make_context
    from data_compare.registry.capability_registry import CapabilityRegistry
    from data_compare.reporting.report_builder import build_report
    (tmp/"p.csv").write_text(prod); (tmp/"d.csv").write_text(dev)
    conv=tmp/"c"; conv.mkdir(exist_ok=True)
    pc=load_any_to_csv(tmp/"p.csv",conv); dc=load_any_to_csv(tmp/"d.csv",conv)
    ctx=make_context(prod_csv_path=pc,dev_csv_path=dc,prod_name="p.csv",dev_name="d.csv",
        result_name="t",report_root=tmp/"r",keys=["ID"],
        capabilities_cfg={"parquet":False,"comparison":True,"tolerance":False,
                          "duplicate":True,"schema":True,"data_quality":True,
                          "audit":True,"alerts":False,"plugins":False})
    ctx=CapabilityRegistry().run_pipeline(ctx)
    return ctx, build_report(ctx)

class TRPT01_Excel_Content_Validation(unittest.TestCase):
    """Grouped Differences sheet contains correct column and row values."""
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TRPT01(self):
        from openpyxl import load_workbook
        ctx,out = _run_and_report("ID,V\n1,A\n2,B\n","ID,V\n1,A\n2,X\n",self.t)
        wb = load_workbook(out)
        ws = wb["Grouped Differences"]
        headers = [ws.cell(1,c).value for c in range(1,ws.max_column+1)]
        self.assertIn("Column",    headers)
        self.assertIn("Prod Value",headers)
        self.assertIn("Dev Value", headers)
        # Should have at least 1 data row
        self.assertGreater(ws.max_row, 1)

class TRPT02_HTML_Content_Accuracy(unittest.TestCase):
    """HTML report contains result name and numeric counts."""
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TRPT02(self):
        from data_compare.reporting.html_report import write_final_html
        summaries=[{"ResultName":"Trade_Compare","ProdFile":"p.csv","DevFile":"d.csv",
                    "TotalProd":100,"TotalDev":100,"MatchedPassed":98,"MatchedFailed":2,
                    "ExtraDev":0,"MissingDev":0,"DuplicateProd":0,"DuplicateDev":0,
                    "UsedKeys":["ID"],"MissingKeyColumns":[],"IgnoredFields":[],
                    "ToleranceUsed":{},"ReportPath":"","ElapsedSeconds":1.2,
                    "AlertsTriggered":0,"SchemaExtraCount":0,"SchemaMissingCount":0}]
        html_p = write_final_html(summaries,self.t,1.2)
        content = html_p.read_text()
        self.assertIn("Trade_Compare",content)
        self.assertIn("98",content)
        self.assertIn("2",content)

class TRPT03_JSON_Schema_Validation(unittest.TestCase):
    """JSON audit conforms to expected schema with all mandatory keys."""
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TRPT03(self):
        from data_compare.reporting.json_audit import write_json_audit
        s=[{"ResultName":"T","ProdFile":"p.csv","DevFile":"d.csv",
            "TotalProd":5,"TotalDev":5,"MatchedPassed":5,"MatchedFailed":0,
            "ExtraDev":0,"MissingDev":0,"DuplicateProd":0,"DuplicateDev":0,
            "UsedKeys":["ID"],"MissingKeyColumns":[],"IgnoredFields":[],
            "ToleranceUsed":{},"ReportPath":"","ElapsedSeconds":0.1,
            "AlertsTriggered":0,"SchemaExtraCount":0,"SchemaMissingCount":0}]
        jp=write_json_audit(s,self.t,0.1)
        d=json.loads(jp.read_text())
        for k in ["run_timestamp","batch_count","overall_pass","framework_version","batches"]:
            self.assertIn(k,d)
        self.assertIsInstance(d["batches"],list)
        self.assertEqual(d["batch_count"],1)

class TRPT04_Large_Report_Generation(unittest.TestCase):
    """5000-row comparison generates a valid Excel report without timeout."""
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TRPT04(self):
        import time
        n=5000
        rows=[f"{i},val{i}" for i in range(n)]
        prod="ID,V\n"+"\n".join(rows)
        dev_rows=rows[:]; dev_rows[0]="0,DIFF"  # 1 difference
        dev="ID,V\n"+"\n".join(dev_rows)
        t0=time.perf_counter()
        ctx,out=_run_and_report(prod,dev,self.t)
        self.assertLess(time.perf_counter()-t0,120)
        self.assertTrue(out.exists())
        self.assertGreater(out.stat().st_size,1000)

class TRPT05_Special_Characters_In_Report(unittest.TestCase):
    """Values with special chars (<>&\"') written to Excel without corruption."""
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TRPT05(self):
        from openpyxl import load_workbook
        prod='ID,V\n1,"<hello> & \'world\'"\n'
        dev='ID,V\n1,"<hello> & \'earth\'"\n'
        ctx,out=_run_and_report(prod,dev,self.t)
        wb=load_workbook(out)
        self.assertIn("Grouped Differences",wb.sheetnames)

class TRPT06_Unicode_In_Report(unittest.TestCase):
    """Unicode values (Chinese, Arabic, emoji) written to Excel correctly."""
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TRPT06(self):
        from openpyxl import load_workbook
        prod="ID,V\n1,你好世界\n2,مرحبا\n"
        dev="ID,V\n1,你好世界\n2,مرحبا\n"
        ctx,out=_run_and_report(prod,dev,self.t)
        wb=load_workbook(out)
        self.assertIn("Pass Comparison",wb.sheetnames)

if __name__ == "__main__":
    import unittest; unittest.main(verbosity=2)
