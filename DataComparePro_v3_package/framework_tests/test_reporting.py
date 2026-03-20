# -*- coding: utf-8 -*-
"""
tests/test_reporting.py
────────────────────────
TR-01  Excel report generated and exists on disk
TR-02  HTML report generated and non-empty
TR-03  JSON audit generated with correct structure
TR-04  JSON overall_pass=True when no failures
TR-05  JSON overall_pass=False when failures present
TR-06  Excel has all mandatory sheets
TR-07  Grouped Differences sheet has correct columns
TR-08  Count Differences sheet has Side column
TR-09  Pass Comparison sheet has CompareResult_ columns
TR-10  Settings sheet has Used Keys row
TR-11  Data Quality Report sheet present when DQ enabled
TR-12  Audit Trail sheet present when audit enabled
TR-13  Alerts sheet present when alerts triggered
TR-14  build_report returns None gracefully when no report_root
TR-15  write_workbook_once writes all provided sheets
"""
from __future__ import annotations
import sys, os, tempfile, json, unittest
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from pathlib import Path
import pandas as pd
from data_compare.context.run_context import make_context
from data_compare.registry.capability_registry import CapabilityRegistry
from data_compare.reporting.report_builder import build_report
from data_compare.reporting.json_audit import write_json_audit
from data_compare.reporting.html_report import write_final_html

def _run_full(prod_text, dev_text, tmp, keys=None):
    from data_compare.loaders.file_loader import load_any_to_csv
    (tmp/"p.csv").write_text(prod_text); (tmp/"d.csv").write_text(dev_text)
    conv=tmp/"c"; conv.mkdir(exist_ok=True)
    pc=load_any_to_csv(tmp/"p.csv",conv); dc=load_any_to_csv(tmp/"d.csv",conv)
    ctx = make_context(prod_csv_path=pc,dev_csv_path=dc,prod_name="p.csv",dev_name="d.csv",
        result_name="report_test",report_root=tmp/"reports",
        keys=keys or ["ID"],
        capabilities_cfg={"parquet":False,"comparison":True,"tolerance":False,"duplicate":True,
                          "schema":True,"data_quality":True,"audit":True,"alerts":True,"plugins":False},
        alert_rules=[{"metric":"matched_failed","operator":">","threshold":0,"level":"WARNING","message":"diffs"}])
    return CapabilityRegistry().run_pipeline(ctx)

class TR01_Excel_Exists(unittest.TestCase):
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TR01(self):
        ctx = _run_full("ID,V\n1,A\n2,B\n","ID,V\n1,A\n2,X\n", self.t)
        out = build_report(ctx)
        self.assertIsNotNone(out); self.assertTrue(out.exists())

class TR02_HTML_NonEmpty(unittest.TestCase):
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TR02(self):
        summaries = [{"ResultName":"t","ProdFile":"p.csv","DevFile":"d.csv",
                      "TotalProd":2,"TotalDev":2,"MatchedPassed":1,"MatchedFailed":1,
                      "ExtraDev":0,"MissingDev":0,"DuplicateProd":0,"DuplicateDev":0,
                      "UsedKeys":["ID"],"MissingKeyColumns":[],"IgnoredFields":[],
                      "ToleranceUsed":{},"ReportPath":"","ElapsedSeconds":0.1,
                      "AlertsTriggered":1,"SchemaExtraCount":0,"SchemaMissingCount":0}]
        html_p = write_final_html(summaries, self.t, 0.1)
        self.assertTrue(html_p.exists()); self.assertGreater(html_p.stat().st_size, 500)

class TR03_JSON_Structure(unittest.TestCase):
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TR03(self):
        summaries = [{"ResultName":"t","ProdFile":"p.csv","DevFile":"d.csv",
                      "TotalProd":2,"TotalDev":2,"MatchedPassed":2,"MatchedFailed":0,
                      "ExtraDev":0,"MissingDev":0,"DuplicateProd":0,"DuplicateDev":0,
                      "UsedKeys":["ID"],"MissingKeyColumns":[],"IgnoredFields":[],
                      "ToleranceUsed":{},"ReportPath":"","ElapsedSeconds":0.1,
                      "AlertsTriggered":0,"SchemaExtraCount":0,"SchemaMissingCount":0}]
        jp = write_json_audit(summaries, self.t, 0.1)
        self.assertTrue(jp.exists())
        d = json.loads(jp.read_text())
        for k in ["run_timestamp","batch_count","overall_pass","framework_version","batches"]:
            self.assertIn(k, d)

class TR04_JSON_Pass(unittest.TestCase):
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TR04(self):
        s = [{"ResultName":"t","ProdFile":"p.csv","DevFile":"d.csv",
              "TotalProd":1,"TotalDev":1,"MatchedPassed":1,"MatchedFailed":0,
              "ExtraDev":0,"MissingDev":0,"DuplicateProd":0,"DuplicateDev":0,
              "UsedKeys":["ID"],"MissingKeyColumns":[],"IgnoredFields":[],
              "ToleranceUsed":{},"ReportPath":"","ElapsedSeconds":0.1,
              "AlertsTriggered":0,"SchemaExtraCount":0,"SchemaMissingCount":0}]
        d = json.loads(write_json_audit(s,self.t,0.1).read_text())
        self.assertTrue(d["overall_pass"])

class TR05_JSON_Fail(unittest.TestCase):
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TR05(self):
        s = [{"ResultName":"t","ProdFile":"p.csv","DevFile":"d.csv",
              "TotalProd":2,"TotalDev":2,"MatchedPassed":1,"MatchedFailed":1,
              "ExtraDev":0,"MissingDev":0,"DuplicateProd":0,"DuplicateDev":0,
              "UsedKeys":["ID"],"MissingKeyColumns":[],"IgnoredFields":[],
              "ToleranceUsed":{},"ReportPath":"","ElapsedSeconds":0.1,
              "AlertsTriggered":0,"SchemaExtraCount":0,"SchemaMissingCount":0}]
        d = json.loads(write_json_audit(s,self.t,0.1).read_text())
        self.assertFalse(d["overall_pass"])

class TR06_All_Sheets(unittest.TestCase):
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TR06(self):
        from openpyxl import load_workbook
        ctx = _run_full("ID,V\n1,A\n","ID,V\n1,B\n", self.t)
        wb = load_workbook(build_report(ctx))
        for s in ["Grouped Differences","Count Differences","Pass Comparison",
                  "Duplicates Records","Schema Differences","Settings",
                  "Data Quality Report","Audit Trail"]:
            self.assertIn(s, wb.sheetnames, f"Missing: {s}")

class TR07_GroupedDiff_Columns(unittest.TestCase):
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TR07(self):
        ctx = _run_full("ID,V\n1,A\n","ID,V\n1,B\n", self.t)
        gd = ctx["results"]["grouped_differences"]
        self.assertIsNotNone(gd); self.assertFalse(gd.empty)
        for col in ["Column","Prod Value","Dev Value"]:
            self.assertIn(col, gd.columns)

class TR08_CountDiff_Side(unittest.TestCase):
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TR08(self):
        ctx = _run_full("ID,V\n1,A\n2,B\n","ID,V\n1,A\n3,C\n", self.t)
        me = ctx["results"]["missing_extra"]
        self.assertFalse(me.empty); self.assertIn("Side", me.columns)

class TR09_PassReport_Columns(unittest.TestCase):
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TR09(self):
        ctx = _run_full("ID,V\n1,A\n","ID,V\n1,B\n", self.t)
        pr = ctx["results"]["pass_report"]
        self.assertIsNotNone(pr)
        cr_cols = [c for c in pr.columns if c.startswith("CompareResult_")]
        self.assertGreater(len(cr_cols), 0)

class TR10_Settings_Keys(unittest.TestCase):
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TR10(self):
        from openpyxl import load_workbook
        from data_compare.reporting.report_builder import build_settings_df
        ctx = _run_full("ID,V\n1,A\n","ID,V\n1,A\n", self.t)
        build_report(ctx)
        settings = build_settings_df(ctx)
        self.assertIn("Used Keys", settings["Setting"].tolist())

class TR11_DQ_Sheet(unittest.TestCase):
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TR11(self):
        from openpyxl import load_workbook
        ctx = _run_full("ID,V\n1,A\n2,B\n","ID,V\n1,A\n2,B\n", self.t)
        wb = load_workbook(build_report(ctx))
        self.assertIn("Data Quality Report", wb.sheetnames)

class TR12_Audit_Sheet(unittest.TestCase):
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TR12(self):
        from openpyxl import load_workbook
        ctx = _run_full("ID,V\n1,A\n","ID,V\n1,A\n", self.t)
        wb = load_workbook(build_report(ctx))
        self.assertIn("Audit Trail", wb.sheetnames)

class TR13_Alerts_Sheet(unittest.TestCase):
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TR13(self):
        from openpyxl import load_workbook
        ctx = _run_full("ID,V\n1,A\n","ID,V\n1,B\n", self.t)
        if ctx["alerts"]["triggered"]:
            wb = load_workbook(build_report(ctx))
            self.assertIn("Alerts", wb.sheetnames)

class TR14_NoReportRoot(unittest.TestCase):
    def test_TR14(self):
        ctx = make_context()
        ctx["config"]["report_root"] = None
        result = build_report(ctx)
        self.assertIsNone(result)

class TR15_WriteWorkbook(unittest.TestCase):
    def setUp(self): self._d=tempfile.TemporaryDirectory(); self.t=Path(self._d.name)
    def tearDown(self): self._d.cleanup()
    def test_TR15(self):
        from data_compare.reporting.excel_report import write_workbook_once
        from openpyxl import load_workbook
        sheets = {"Sheet1": pd.DataFrame({"A":[1,2],"B":[3,4]}),
                  "Sheet2": pd.DataFrame({"X":["a","b"]})}
        out = self.t/"out.xlsx"; write_workbook_once(sheets, out)
        self.assertTrue(out.exists())
        wb = load_workbook(out)
        self.assertIn("Sheet1", wb.sheetnames); self.assertIn("Sheet2", wb.sheetnames)

if __name__ == "__main__":
    import unittest; unittest.main(verbosity=2)
